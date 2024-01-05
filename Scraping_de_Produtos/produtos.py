from bs4 import BeautifulSoup
import requests
from tkinter import messagebox
import openpyxl
import os

class Produtos:
    def __init__(self, categoria):
        self.categoria = categoria
        self.items = []
        self.dados_produtos=[]
        self.obter_dados()
        self.listar_produtos()
        
        
    def obter_dados(self):
        response = requests.get(self.categoria)
        soup = BeautifulSoup(response.text,'html.parser')
        self.items = soup.select('.item')
        
    def listar_produtos(self):
        #print("######################## Produtos Listados ########################")
        for item in self.items:
            product_name = item.select_one('.item-info-container a').text.strip().upper()
            preco = item.select_one('.item-price').text.strip()
            links = item.select_one('.item-image-container a')['href']
            status = "Esgotado" if "Esgotado" in item.text else "5% OFF"
            self.dados_produtos.append([product_name,preco,status,links])
            
    def obter_nome_planilha(self):
        if self.categoria == "https://meufornecedor.com/":
            return "Em Destaque"
        elif self.categoria == "https://www.meufornecedor.com/produtos/":
            return "Todos Os Produtos"
        else:
            return "Produtos Originais"
    
    def gerar_planilha_xlsx(self):
        nome_planilha = self.obter_nome_planilha()
        caminho_arquivo = os.path.join(os.path.expanduser("~/Documents"),"produtos_para_anunciar.xlsx")
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            
        # Adiciona uma nova sheet ou usa uma existente
        if nome_planilha not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=nome_planilha)
            headers = ["Produto", "Preço", "Status", "Link"]
            sheet.append(headers)
        else:
            sheet = workbook[nome_planilha]

        for dado in self.dados_produtos:
            sheet.append(dado)
        workbook.save(caminho_arquivo)
        messagebox.showinfo("Concluído","Um arquivo Excel foi gerado na sua pasta Documents com as informações extraídas", icon="info")
        
